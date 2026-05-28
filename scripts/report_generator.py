"""
report_generator.py
===================
Automated Excel report generator.
Copy into any project. Fill in YOUR DATA HERE sections.

RUN:
    python scripts/report_generator.py
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

# ── CONFIG ────────────────────────────────────────────────────
REPORT_NAME  = 'YOUR_REPORT_NAME'   # ← change this
FILTER_YEAR  = None                  # ← set to 2024 or None for all
HEADER_COLOR = '1F3864'             # ← dark navy — change if needed
OUTPUT_DIR   = 'reports'


# ── STEP 1: CONNECT & PULL ────────────────────────────────────
def pull_data():
    from scripts.db_connect import pull_table, query

    print("Pulling data...")

    # ── YOUR DATA HERE ────────────────────────────────────────
    df_main = pull_table('your_main_table')
    # df_ref  = pull_table('your_reference_table')
    # ──────────────────────────────────────────────────────────

    print(f"Pulled {len(df_main):,} rows")
    return df_main


# ── STEP 2: CLEAN ─────────────────────────────────────────────
def clean_data(df):
    from scripts.cleaning_utils import run_clean_pipeline

    # ── YOUR DATA HERE ────────────────────────────────────────
    config = {
        'clean_column_names': True,
        'text_columns':   [],        # ← add your text columns
        'currency_cols':  [],        # ← add your money columns
        'date_cols':      [],        # ← add your date columns
        'date_parts_col': None,      # ← set to your main date col
        'required_cols':  [],        # ← columns that must not be null
    }
    # ──────────────────────────────────────────────────────────

    df = run_clean_pipeline(df, config)

    if FILTER_YEAR and 'year' in df.columns:
        df = df[df['year'] == FILTER_YEAR]
        print(f"Filtered to {FILTER_YEAR}: {len(df):,} rows")

    return df


# ── STEP 3: CALCULATE KPIs ────────────────────────────────────
def calculate_kpis(df):
    print("Calculating KPIs...")

    # ── YOUR DATA HERE ────────────────────────────────────────
    kpis = {
        'KPI 1 Label': f"{df['your_col'].sum():,.2f}",  # ← change
        'KPI 2 Label': f"{len(df):,}",
        'KPI 3 Label': f"{df['your_col'].mean():,.2f}",
    }
    # ──────────────────────────────────────────────────────────

    summary = pd.DataFrame(list(kpis.items()), columns=['KPI', 'Value'])
    return summary


# ── STEP 4: BUILD SUMMARIES ───────────────────────────────────
def build_summaries(df):
    print("Building summaries...")

    summaries = {}

    # ── YOUR DATA HERE ────────────────────────────────────────
    # summaries['By Category'] = df.groupby('your_col').agg(
    #     total = ('amount_col', 'sum'),
    #     count = ('id_col', 'count'),
    # ).round(2).reset_index().sort_values('total', ascending=False)
    # ──────────────────────────────────────────────────────────

    return summaries


# ── STEP 5: EXPORT ────────────────────────────────────────────
def export_report(summary_kpis, summaries, df_raw):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    report_date = datetime.now().strftime('%Y-%m-%d_%H-%M')
    filename = f"{OUTPUT_DIR}/{REPORT_NAME}_{report_date}.xlsx"

    print("Building Excel report...")

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        summary_kpis.to_excel(writer, sheet_name='Executive Summary', index=False)
        for sheet_name, df_summary in summaries.items():
            df_summary.to_excel(writer, sheet_name=sheet_name, index=False)
        df_raw.to_excel(writer, sheet_name='Raw Data', index=False)

    # Format
    wb = load_workbook(filename)
    header_font  = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill  = PatternFill('solid', fgColor=HEADER_COLOR)
    header_align = Alignment(horizontal='center', vertical='center')
    data_font    = Font(name='Arial', size=10)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.row_dimensions[1].height = 22
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
        for col in ws.columns:
            max_len = max(
                (len(str(c.value)) for c in col if c.value), default=10
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)
        ws.freeze_panes = 'A2'

    wb.save(filename)
    print(f"Report saved: {filename}")
    return filename


# ── MAIN ──────────────────────────────────────────────────────
if __name__ == '__main__':
    df_raw      = pull_data()
    df_clean    = clean_data(df_raw.copy())
    kpi_summary = calculate_kpis(df_clean)
    summaries   = build_summaries(df_clean)
    export_report(kpi_summary, summaries, df_clean)
    print("Done!")
